"""Module which handles reading and writing of excel files"""

import os
import openpyxl as opxl
from openpyxl.styles import PatternFill
import python_calamine
from data_structures import Participant, Group, Iteration, Assignment


class Reader:
    """class that reads excel sheets and turns the table into a list of participants using python_calamine"""

    @staticmethod
    def read(filepath: os.PathLike) -> list[Participant]:
        """Reads an excel file on the speciefied location and creates a list of participants from it.

        :param filepath: The path to the file to be read
        :return: The list of participants with initialized attributes"""
        workbook = python_calamine.CalamineWorkbook.from_path(filepath)
        rows = iter(workbook.get_sheet_by_index(0).to_python())

        headers = list(map(str, next(rows)))

        participant_list: list[Participant] = []
        j: int = 0

        for row in rows:
            p: Participant = Participant(j)

            # All dates will be represented in the following form: 2025-01-31
            for i, header in enumerate(headers):
                if isinstance(row[i], float) and row[i].is_integer():
                    p.attributes[header] = f"{row[i]:.0f}"
                else:
                    p.attributes[header] = str(row[i])

            participant_list.append(p)
            j += 1

        return participant_list


class Writer:
    """Class that writes excel sheet and turns calculated groups in an understandable format.
    Only meant to be used once. Object of Writer should only be present when writing file
    """

    __row_index: int = 1
    __filepath: os.PathLike

    # Colors for coloring the first cell for better understandability
    __fill_colors: tuple[PatternFill] = (
        PatternFill(start_color="00CCFFCC", fill_type="solid"),  # green
        PatternFill(start_color="00CC99FF", fill_type="solid"),  # violet
    )

    def __init__(self, filepath: os.PathLike) -> None:
        self.__filepath = filepath

    def __write_header(
        self, iteration_number: int, attribute_list: list[str], ws
    ) -> None:
        """This function writes the header for an iteration with the iteration number and
        header row(group, list of attributes).

        :param iteration_number: number of the iteration to be written
        :param attribute_list: list of attribute names of participants
        :param row_index: row index to be written to
        :param ws: worksheet to be written on
        """

        ws.cell(self.__row_index, 1).value = f"Iteration {iteration_number}:"
        self.__row_index += 1
        ws.cell(self.__row_index, 1).value = "GroupNr"

        for i, attribute in enumerate(attribute_list):
            ws.cell(self.__row_index, 2 + i).value = attribute

    def __write_participant(
        self,
        participant: Participant,
        group_number: int,
        attribute_list: list[str],
        ws,
    ) -> None:
        """This function writes a participant with its group number (colored background) and
        its attribute values.

        :param participant: participant to be written
        :param group_number: group number of the participant
        :param attribute_list: list of attribute names of participants
        :param ws: worksheet to be written to
        """

        ws.cell(self.__row_index, 1).fill = self.__fill_colors[
            group_number % len(self.__fill_colors)
        ]

        ws.cell(self.__row_index, 1).value = group_number

        for i, attribute in enumerate(attribute_list):
            ws.cell(self.__row_index, 2 + i).value = participant.get_attribute(
                attribute
            )

        self.__row_index += 1

    def __write_group(
        self,
        group: Group,
        group_number: int,
        attribute_list: list[str],
        ws,
    ) -> None:
        """This function writes all members of a grounp to the worksheet.

        :param group: group to be written
        :param group_number: group number of the group
        :param attribute_list: list of attribute names of participants
        :param ws: worksheet to be written to
        """

        for participant in iter(group):
            self.__write_participant(participant, group_number, attribute_list, ws)

    def __write_iteration(
        self,
        iteration: Iteration,
        iteration_number: int,
        attribute_list: list[str],
        ws,
    ) -> None:
        """This function writes the iteration header and paricipants of all groups to the worksheet.

        :param iteration: iteration to be written
        :param iteration_number: number of the iteration that is written
        :param attribute_list: list of attribute names of participants
        :param ws: worksheet to be written to
        """

        # write iteration
        self.__write_header(iteration_number, attribute_list, ws)

        self.__row_index += 1

        for i, group in enumerate(iteration):
            self.__write_group(group, i + 1, attribute_list, ws)

        self.__row_index += 2  # add 2 extra empty rows for better readability

    def __write_assignment(self, assignment: Assignment, ws) -> None:
        """This function writes an assignment with all its iteration.

        :param assignment: assignment to be printed
        :param row_index: row index to be written to
        :param ws: worksheet to be written to
        """

        attribute_list: list[str] = list(next(iter(assignment[0][0])).attributes.keys())

        for i, iteration in enumerate(assignment):
            self.__write_iteration(iteration, i + 1, attribute_list, ws)

    def write_file(self, assignment: Assignment) -> None:
        """This method is used to write the excel sheet to the path that is set containing
        the iterations with its groups.

        :param assignment: The assignment to be written to the excel file
        """

        self.__row_index: int = 1

        wb = opxl.Workbook()
        ws = wb.worksheets[0]

        self.__write_assignment(assignment, ws)

        wb.save(self.__filepath)
